import urllib.request
import src.db_operations as db
import base64
import requests
import re
import d20
from bs4 import BeautifulSoup as bs4
from openpyxl.utils.cell import get_column_letter



def add_values_to_formula_string(dictionary:dict, string:str):
    ###sort dictionary so longest keys go first (in case a dictionary key is a substring of another key)
    sorted_keys = sorted(dictionary.keys(), key=len, reverse=True)
    sorted_dict = dict(zip(sorted_keys, [dictionary[key] for key in sorted_keys]))

    ###regex to not match items placed by matches in previous iterations
    for key, value in sorted_dict.items():
        regex = str(key) + '(?![\w]*[\]])'
        replacement = str(value) + "[" + str(key) + "]"
        string = re.sub(pattern=regex, repl=replacement, string=string)

    return string

def is_link_image (link:str):
    try:
        image_formats = ("image/png", "image/jpeg", "image/webp")
        site = urllib.request.urlopen(link)
        meta = site.info()
        print(meta["content-type"])
        if meta["content-type"] in image_formats:
            return True
        else:
            return False
    except Exception as e:
        print(e)
        return str(e)
    
def get_image_token(url:str):
    url64 = "https://token.otfbm.io/meta/" + encode_url_base64(url)
    response = requests.get(url64)
    if response.status_code == 200:
        soup = bs4(response.text, "html.parser")
        return soup.find("body").text.strip()
    return False
    
def encode_url_base64(url:str):
    return str(base64.standard_b64encode(url.encode("utf-8")), "utf-8")

def test_formula(rules, formula:str):
    dictionary = {}
    for rule in rules:
        dictionary[rule] = 1
    
    try:
        d20.roll(add_values_to_formula_string(dictionary, formula)).total
    except d20.errors.RollSyntaxError:
        return "The formula can't be evaluated. Either there is an incorrect variable or the formula itself is incorrect."
    
    dictionary = {}
    for rule in rules:
        dictionary[rule] = 0
    
    try:
        d20.roll(add_values_to_formula_string(dictionary, formula)).total
    except d20.errors.RollValueError:
        return "The formula may return a division by zero error."
    
    return True

def map_url_constructor(map, entities=None, ability=None, ability_coords=None):
    url = "https://otfbm.io/" + map["length"] + "x" + map["height"] + "/@dc" + map["grid_size"]

    for entity in entities:
        if entity.player_id == 0:
            color = "r"
        else:
            color = "g"

        url = url + "/" + get_column_letter(entity.x) + str(entity.y) + color + "-" + entity.name
        if entity.token != "":
            url = url + "~" + entity.token
    

    if ability != None:
        match ability.type:
            case "single":
                url = url + "/*c2b" + get_column_letter(ability.x) + str(ability.y)
            case "line":
                url = url + "/*l" + str(int(ability.range) * 5) + ",1" + get_column_letter(ability_coords[0]) + str(ability_coords[1]) + get_column_letter(ability.x) + str(ability.y)
                url = url + "/*c2b" + get_column_letter(ability.x) + str(ability.y)
            case "area":
                url = url + "/*c" + str(int(ability.range) * 5) + "b" + get_column_letter(ability_coords[0]) + str(ability_coords[1])
            case "cone":
                url = url + "/t" + str(int(ability.range) * 5) + "b" + get_column_letter(ability_coords[0]) + str(ability_coords[1]) + get_column_letter(ability.x) + str(ability.y)
                url = url + "/*c2b" + get_column_letter(ability.x) + str(ability.y)


    url = url+ "/?bg=" + map["image"]
    return url
