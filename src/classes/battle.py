import src.db_operations as db
import src.classes.creature
import src.classes.character
import src.classes.effect
from skimage.draw import (line, disk)
from openpyxl.utils.cell import get_column_letter
import math
import numpy as np



class Battle():

    def __init__(self, encounter, server_id, characters):
        self.server_id = server_id
        self.map = db.get_object("maps", encounter["map"], server_id)
        self.map_length = self.map["length"]
        self.map_height = self.map["height"]
        self.name = encounter["name"]
        self.description = encounter["description"]
        self.entities = []
        self.turn = 1
        self.char_turn = 0
        self.ability_labels = []
        self.ability_pages = 1

        self.thread = None
        self.over = False

        self.current_char_actual_x = 0
        self.current_char_actual_y = 0
        self.current_char_actual_moves = 0
        self.current_char_moved = False
        self.current_char_ability_used = False
        self.ability_mode = False
        self.active_ability = None

        for creature in encounter["creatures"]:
            creature_data = db.get_object("creatures", creature["name"], self.server_id)
            self.entities.append(src.classes.creature.Creature(creature_data, creature["length_coord"], creature["height_coord"], self.server_id, self.thread))

        for character in characters:
            character_data = db.get_object("characters", character["name"], self.server_id)
            self.entities.append(src.classes.character.Character(character_data, character["length_coord"], character["height_coord"], self.server_id, self.thread))

        self.entities.sort(reverse=True, key=lambda x: x.initiative)






    def get_current_entity(self):
        return self.entities[self.char_turn]
    

    async def initialize_char_turn(self):
        
        self.current_char_actual_x = self.get_current_entity().x
        self.current_char_actual_y = self.get_current_entity().y
        self.current_char_moved = False
        self.current_char_ability_used = False
        self.ability_mode = False
        await self.get_current_entity().initialize_turn(self.thread)
        self.current_char_actual_moves = self.get_current_entity().speed

        self.ability_labels.clear()
        for ability in self.get_current_entity().abilities:
            self.ability_labels.append(ability.name)
        self.ability_pages = math.ceil(len(self.ability_labels) / 8)

        ##### Code to run if trying to initialize a defeated character
        if self.get_current_entity().defeated == True:
            await self.thread.send(f"{self.get_current_entity().name} has been defeated. Removing from the board")
            self.entities.remove(self.get_current_entity())

            entity_count = len(self.entities)
            if self.char_turn >= entity_count:
                self.char_turn = 0
                self.turn = self.turn + 1

            await self.initialize_char_turn()



    async def initiative_logs(self):
        msg = ""
        for entity in self.entities:
            msg = msg + entity.init_log_rez + "\n"
        await self.thread.send(msg)

    async def confirm_move(self):
        if not self.ability_mode:
            msg = f"{self.get_current_entity().name} moved from {get_column_letter(self.current_char_actual_x)}{self.current_char_actual_y} to {get_column_letter(self.get_current_entity().x)}{self.get_current_entity().y}"
            self.current_char_actual_x = self.get_current_entity().x
            self.current_char_actual_y = self.get_current_entity().y
            self.get_current_entity().speed = 0
            self.current_char_moved = True   

            await self.thread.send(msg)    
        else:
            msg = f"{self.get_current_entity().name} cast {self.active_ability.name} on {get_column_letter(self.active_ability.x)}{self.active_ability.y}"
            await self.thread.send(msg)

            self.ability_mode = False
            self.current_char_ability_used = True
            await self.add_self_effects()
            await self.add_effects(self.get_entities_hit_by_ability())
            
        


    async def end_turn(self):
        await self.thread.send(f"{self.get_current_entity().name} ends their turn.")
        entity_count = len(self.entities)
        if self.char_turn + 1 == entity_count:
            self.char_turn = 0
            self.turn = self.turn + 1

            await self.thread.send(f"Starting turn: {self.turn}")
        else:
            self.char_turn = self.char_turn + 1
        await self.initialize_char_turn()

    async def add_self_effects(self):
        for self_effect in self.active_ability.self_effects:
            effect_data = db.get_object("effects", self_effect, self.server_id )
            self.get_current_entity().effects.append(src.classes.effect.Effect(effect_data, self.server_id, self.get_current_entity().statistics))
            msg = f"Applied {self_effect} to caster {self.get_current_entity().name}"
            await self.thread.send(msg)

    async def add_effects(self, targets):
        for effect in self.active_ability.effects:
            for target in targets:
                effect_data = db.get_object("effects", effect, self.server_id )
                target.effects.append(src.classes.effect.Effect(effect_data, self.server_id, self.get_current_entity().statistics))
                msg = f"Applied {effect} to target {target.name}"
                await self.thread.send(msg)

    
    def get_entities_hit_by_ability(self):
        match self.active_ability.type:
            case "single":
                for entity in self.entities:
                    if entity.x == self.active_ability.x and entity.y == self.active_ability.y:
                        return [entity]
            case "line":
                #### extend line end coordinates to be farther than the range
                while math.sqrt(math.pow((self.get_current_entity().x - self.active_ability.x), 2) + math.pow((self.get_current_entity().y - self.active_ability.y), 2)) < self.active_ability.range:
                    if self.get_current_entity().x > self.active_ability.x:
                        self.active_ability.x = self.active_ability.x - (self.get_current_entity().x - self.active_ability.x)
                    else:
                        self.active_ability.x = self.active_ability.x + (self.active_ability.x - self.get_current_entity().x)

                    if self.get_current_entity().y > self.active_ability.y:
                        self.active_ability.y = self.active_ability.y - (self.get_current_entity().y - self.active_ability.y)
                    else:
                        self.active_ability.y = self.active_ability.y + (self.active_ability.y - self.get_current_entity().y)

                    if self.get_current_entity().x == self.active_ability.x and self.get_current_entity().y == self.active_ability.y:
                        break


                line_coord = line(self.get_current_entity().x, self.get_current_entity().y, self.active_ability.x, self.active_ability.y)
                rez = []

                for row, column in zip(line_coord[0], line_coord[1]):
                    ### check if the distance is more than range and break if that is the case
                    if math.sqrt(math.pow((self.get_current_entity().x - row), 2) + math.pow((self.get_current_entity().y - column), 2)) > int(self.active_ability.range):
                        break
                    
                    for entity in self.entities:
                        if (entity.x == row and entity.y == column) and not (self.get_current_entity().x == row and self.get_current_entity().y == column):
                            rez.append(entity)
                            break
                return rez
                
            case "area":
                disk_coord = disk((self.get_current_entity().x, self.get_current_entity().y), radius= float(self.active_ability.range + 0.5))
                rez = []

                for row, column in zip(disk_coord[0], disk_coord[1]):
                    for entity in self.entities:
                        if (entity.x == row and entity.y == column) and not (self.get_current_entity().x == row and self.get_current_entity().y == column):
                            rez.append(entity)
                            break
                
                return rez
